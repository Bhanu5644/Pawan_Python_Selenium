import openpyxl

#Multiple data
file="C:/Users/BHANU/OneDrive/Writing Data in Excel.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Munish"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=124
sheet.cell(2,2).value="Sweety"
sheet.cell(2,3).value="Nurse"

sheet.cell(3,1).value=125
sheet.cell(3,2).value="Aahana"
sheet.cell(3,3).value="Student"
workbook.save(file)  # save the file after entering the data