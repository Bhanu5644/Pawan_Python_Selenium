import openpyxl

#File-->Workbook--->Sheets--->Rows---->Cells
File ="C:/Users/BHANU/OneDrive/Data.xlsx"
workbook = openpyxl.load_workbook(File)
Sheet = workbook["Sheet1"]
rows = Sheet.max_row #count no. of rows in a excel file
clos = Sheet.max_column #count no. of column in a excel file
for r in range (1,rows+1):
    for c in range (1,clos+1):
        # print(Sheet.cell(r,c).value)
        print(Sheet.cell(r,c).value,end="             ")
    print()

